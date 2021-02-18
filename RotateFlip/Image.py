from PIL import Image
import openpyxl as xl
import numpy as np
import time
import math
import os


class RotateFlip(object):
    def __init__(self, image_name):
        self.image_name = image_name
        if self.__controller__() == 1:
            while True:
                self.image_name = str(input(f"There is no such file named '{self.image_name}'\n"
                                            f"Please enter a valid filename! "))
                if self.__controller__() == 0:
                    break
        self.__load_image__()

    def __controller__(self):
        try:
            Image.open(self.image_name)
            return 0
        except:
            return 1

    def __load_image__(self, ai=False):
        self.image = Image.open(self.image_name)
        self.pixel = self.image.load()
        self.max_x, self.max_y = self.image.size[0], self.image.size[1]
        if ai is True:
            self.pix_matrix = []
            for y in range(self.max_y):
                pix_line = []
                for x in range(self.max_x):
                    pix = (self.pixel[x, y][0],
                           self.pixel[x, y][1],
                           self.pixel[x, y][2])
                    pix_line.append(pix)
                self.pix_matrix.append(pix_line)

    def __save_image__(self, image):
        type_control = type(image) == type(Image.open("sample.png"))
        if type_control is False:
            raise TypeError

        tmp = self.image_name
        tmp2 = tmp
        base = self.image_name[:len(self.image_name)-4]
        extension = self.image_name[-3:]
        counter = 0
        while True:
            if self.__controller__() == 1:
                self.image_name = tmp
                break
            self.image_name = f"{base}_copy_{counter}.{extension}"
            tmp2 = self.image_name
            counter += 1

        image.save(tmp2)

    def __show_image__(self, image):
        pwd = os.getcwd()
        route = f"{pwd}/{self.image_name}"
        image.show(route)

    def get_image_name(self):
        return self.image_name

    def set_image_name(self, new_name):
        self.image_name = new_name

    def help(self):
        print("Under development!")

    def __str__(self):
        return f"This is a powerful module for rotating and flipping PNG and JPG images.\n" \
               f"Your image's name is '{self.image_name}'.\n" \
               f"You can reach help module with typing RotateFlip.help()\n" \
               f"Thank you for using RotateFlip!"

    def xl_export(self):
        start_time = time.time()

        wb = xl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(f"{self.image_name}_data.xlsx")

        wb = xl.load_workbook(f"{self.image_name}_data.xlsx")
        sheet = wb["Sheet1"]

        self.__load_image__(ai=True)

        for row in range(len(self.pix_matrix)):
            for column in range(len(self.pix_matrix[row])):
                data = self.pix_matrix[row][column]
                data = f"({data[0]},{data[1]},{data[2]})"
                data_cell = sheet.cell(row+1, column+1)
                data_cell.value = data

        wb.save(f"{self.image_name}_data.xlsx")

        end_time = time.time()

        print(f"Process finished in {math.ceil(end_time - start_time)} seconds!")

    def txt_export(self):
        start_time = time.time()

        with open(f"{self.image_name}_data.txt", "w") as f:
            self.__load_image__(ai=True)
            output = ""
            for y in range(len(self.pix_matrix)):
                for x in range(len(self.pix_matrix[y])):
                    output += f"{self.pix_matrix[y][x]} "
                output += "\n"
            f.writelines(output)

        end_time = time.time()

        print(f"Process finished in {math.ceil(end_time - start_time)} seconds!")

    def flip_x(self):
        new_image = self.image
        pixel = self.pixel
        start_time = time.time()

        pix_list = []
        for y in range(self.max_y):
            pix_line = []
            for x in range(self.max_x):
                pix = (self.pixel[x, self.max_y-1-y][0],
                       self.pixel[x, self.max_y-1-y][1],
                       self.pixel[x, self.max_y-1-y][2])
                pix_line.append(pix)
            pix_list.append(pix_line)

        for y in range(len(pix_list)):
            for x in range(len(pix_list[y])):
                pixel[x, y] = pix_list[y][x]

        self.__save_image__(new_image)
        end_time = time.time()

        print(f"Process finished in {math.ceil(end_time - start_time)} seconds!\n"
              f"Image Size:{self.max_x}x{self.max_y}\n")

    def flip_y(self):
        new_image = self.image
        pixel = self.pixel
        start_time = time.time()

        pix_list = []
        for y in range(self.max_y):
            pix_line = []
            for x in range(self.max_x):
                pix = (self.pixel[self.max_x-1-x, y][0],
                       self.pixel[self.max_x-1-x, y][1],
                       self.pixel[self.max_x-1-x, y][2])
                pix_line.append(pix)
            pix_list.append(pix_line)

        for y in range(len(pix_list)):
            for x in range(len(pix_list[y])):
                pixel[x, y] = pix_list[y][x]

        self.__save_image__(new_image)
        end_time = time.time()

        print(f"Process finished in {math.ceil(end_time - start_time)} seconds!\n"
              f"Image Size:{self.max_x}x{self.max_y}\n")

    def boarder_scan(self):
        new_image = self.image
        pixel = self.pixel
        start = time.time()

        general = []
        for y in range(self.max_y):
            tmp = []
            for x in range(self.max_x):
                a = [self.pixel[x, y][0],
                     self.pixel[x, y][1],
                     self.pixel[x, y][2]]
                tmp.append(a)
            general.append(tmp)

        end = time.time()

        print(f"Indexing finished in {math.ceil(end - start)} seconds!\n"
              f"Image Size:{self.max_x}x{self.max_y}")

        start = time.time()

        for i in range(len(general)):
            for j in range(len(general[i])):
                try:
                    q = abs(general[i][j][0] - general[i][j - 10][0])
                    w = abs(general[i][j][1] - general[i][j - 10][1])
                    e = abs(general[i][j][2] - general[i][j - 10][2])
                    if (q + w + e) / 3 > 20:
                        for k in range(1):
                            pixel[j + k, i] = (230, 1, 1)
                            pixel[j - k, i] = (230, 1, 1)
                except IndexError:
                    pass

        self.__save_image__(new_image)
        end = time.time()

        print(f"Borders are created in {math.ceil(end - start)} seconds!")

    def rotate(self, angle=0):
        start_time = time.time()

        self.__load_image__(ai=True)

        angle = angle % 360
        if angle < 0:
            angle = 360 - abs(angle)

        h = (self.max_y*math.cos(angle)) + (self.max_x*math.sin(angle))
        w = (self.max_y*math.sin(angle)) + (self.max_x*math.cos(angle))

        new_image = Image.new("RGB", (w, h), color=(220, 220, 220))
        new_pixel = new_image.load()

        self.__save_image__(new_image)
        end_time = time.time()

        print(f"Indexing finished in {math.ceil(end_time - start_time)} seconds!\n"
              f"Image Size:{w}x{h}")

    def sunset_effect(self, percent=50):
        new_image = self.image
        pixel = self.pixel
        percent = percent / 2
        start_time = time.time()

        self.__load_image__(ai=True)

        for y in range(self.max_y):
            for x in range(self.max_x):
                v1, v2, v3 = pixel[x, y][0],\
                             pixel[x, y][1],\
                             pixel[x, y][2]
                ideal = np.average([2*(115-v3), 3*(np.average([v1, v2])-175)])\
                        *percent

                v1, v2, v3 = math.ceil(v1*(100+2*ideal)/100),\
                             math.ceil(v2*(100+ideal)/100),\
                             math.ceil(v3*(100-ideal)/100)

                if v1 > 250 or v1 < 1:
                    if v1 > 250:
                        v1 = 250
                    else:
                        v1 = 1
                if v2 > 250 or v2 < 1:
                    if v2 > 250:
                        v2 = 250
                    else:
                        v2 = 1
                if v3 > 250 or v3 < 1:
                    if v3 > 250:
                        v3 = 250
                    else:
                        v3 = 1

                pixel[x, y] = (v1, v2, v3)

        self.__save_image__(new_image)
        end_time = time.time()

        print(f"Effects are processed in {math.ceil(end_time - start_time)} seconds!")
